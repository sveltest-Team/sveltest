#!/usr/bin/env python
#-*- coding:utf-8 -*-



import os
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

#
# TODO: 未完成
class OpenExecl(object):
    """
    用于对Excel操作、如读写
    """
    def __init__(self,file):
        self.excel = openpyxl.load_workbook(file, data_only=True)
        self.sheet = None
        self.cols = self.active().columns
        self.rows = self.active().rows
        self.filename = self.get_sheets()
        self.sheetnames = None
        self.current_sheet = self.excel.active
        self.count_cols = None
        self.max_row = self.active().max_row #行
        self.max_col = self.active().max_column #列


    def set_active(self,index):
        """
        设置当前的活动表
        :param index: 输入的是一个表格的索引值
        :return:
        """
        self.excel.active = index
        return self.excel.active


    def active(self):
        """

        :return:
        """
        return self.excel.active

    def get_sheets(self):
        return self.excel.sheetnames

    # # 切换指定的工作表
    # def switch(self,name):
    #     ws = self.excel[sheets[n]]

    # def get_cols_all(self):
    #     col = []
    #     for column in self.active().columns:
    #         cols_list = []
    #         for i in column:
    #             cols_list.append(i.value)
    #         col.append(cols_list)
    #     return col

    def get_cols_all(self):
        """

        """
        col = []
        for column in self.active().columns:
            print(column)
            cols_list = []
            for i in column:
                print(i)
                cols_list.append(i.value)
            col.append(cols_list)
        return col

    def cell(self,x,y):
        return self.excel.active.cell(x,y).value

    def get_rows_all(self,change=False):
        """
        按照行进行输出
        change : 是否对数据进行数据类型转换
        :return:
        """
        row = [i.value for rows in self.active().rows  for i in rows]
        row = []
        for rows in self.active().rows:
            row_list = [i.value for i in rows]
            row.append(row_list)
        return row

    def get_index_row(self):pass

    def get_index_col(self):pass

    def get_row(self):
        return self.active().rows

    def set_sheet_name(self,name):
        # 切换指定的工作表
        """
        drs = dr.get_sheets()
        set_sheet_name(drs[0])
        :param name:
        :return:
        """
        return self.excel[name]

    # def set_active(self):



# TODO:第一版基础功能完成 待测试
# guanfl
# v1.0
# 20210619
# def get_script(path=None, set_sheet_name=None):
#     dr = TankOpenExecl(file=r'F:\tank\src\tank\default\自动化.xlsx')
#     # 获取表单的列和行数
#     get_col_row = [dr.max_row,dr.max_col]
#
#     for d in range(len(dr.get_sheets())):
#         # 进行限制没有进行修改工作表名将进行不出来代码生成
#         if dr.get_sheets()[d][0:-1] != "Sheet":
#             # 存储所有数据
#             data_all = []
#
#             dr.set_active(d)
#
#             for c in range(0,get_col_row[0]):
#                 # 存储每一行的数据
#                 cs = []
#                 for r in range(get_col_row[1]):
#                     cs.append(dr.cell(c+1,r+1))
#                 #每一行数据进行存储到所有数据列表中
#                 data_all.append(cs)
#
#             func_list = []
#
#             with open(f'{dr.filename[d]}.py', "w+", encoding="utf-8") as f:
#
#                 for s in range(len(data_all)):
#                     if s !=0:
#                         if data_all[s][0] == "url":
#                             f.write("from selenium import webdriver\n")
#                             f.write("from selenium.webdriver.common.by import By\n")
#                             f.write("import time\n\n\n")
#
#                             f.write(f"class {dr.filename[d]}:\n\n")
#                             f.write(f"    def __init__(self):\n")
#                             url = str(data_all[s][1])
#                             try:
#                                 note = data_all[s][2]
#                                 if note:f.write(f"        #{note}\n")
#                             except:
#                                 pass
#                             f.write(f"        self.url = '{url}'\n")
#
#                         if data_all[s][0] == "open":
#                             driver = data_all[s][1]
#                             try:
#                                 note = data_all[s][2]
#                                 if note:f.write(f"        #{note}\n")
#                             except:
#                                 pass
#
#                             if driver == "chrome" or driver == "Chrome":
#                                 driver = "Chrome"
#                                 f.write(f"        self.driver = webdriver.{driver}()\n")
#
#                             if driver == "ie" or driver == "IE":
#                                 f.write(f"        self.driver = webdriver.Ie()\n")
#                             f.write(f"        self.driver.get(self.url)\n\n")
#
#                         if data_all[s][0] == "func" :
#                             func_list.append(data_all[s][1])
#                             f.write(f"    def test_{data_all[s][1]}(self):\n")
#
#
#                         if data_all[s][0] == "sleep":
#                             try:
#                                 note = data_all[s][2]
#                                 if note:f.write(f"        #{note}\n")
#                             except:
#                                 pass
#                             f.write(f"        time.sleep(4)\n")
#
#                         if data_all[s][0] == "input":
#
#
#                             try:
#                                 note = data_all[s][3]
#                                 if note:f.write(f"        #{note}\n")
#                             except:
#                                 pass
#                                 # upper() #转换成大写
#                             element_by = str(data_all[s][1]).split("=")
#                             f.write(f"        self.driver.find_element(by=By.{element_by[0].upper()},value='{element_by[1]}').send_keys('{data_all[s][2]}')\n")
#
#
#                         if data_all[s][0] == "end":
#
#
#                             try:
#                                 note = data_all[s][3]
#                                 if note:f.write(f"        #{note}\n")
#                             except:
#                                 pass
#
#                             try:
#                                 f.write(f"\n\n\nif __name__ == '__main__':\n    case = {dr.filename[d]}()\n    case.test_{func_list[0]}()\n")
#                             except:
#                                 print("不允许在没有声明方法前进行使用end关键字")
#                             break
#
#




# postman export json analysis
# 解析

